"""
modules/tsi_loader.py
---------------------
Parser for TSI SMPS Excel export files (AIM software output).

TSI SMPS files store data in one sheet per dataset with a fixed layout:

  Block 1 – Instrument metadata
      Rows from the top, each with a descriptor in column A and the
      corresponding value in column B.  Ends just before the scan header.

  Block 2 – Scan metadata  (embedded in the scan rows)
      Starting from the "Scan Number" header row, each scan occupies three
      consecutive rows.  Row +0 contains metadata for that scan (scan index,
      datetime, flow settings, voltage range, statistical summaries, …).

  Block 3 – dN/dlogDp concentration grid  (same row +0 as block 2)
      Immediately after the metadata columns in row +0, columns contain
      dN/dlogDp values, one column per diameter bin.  The diameter midpoints
      (nm) are given in the same columns of the header row.

  Block 4 – Raw data  (rows +1 and +2 of each scan block)
      The header row has "Raw Data - Time (s)" followed by the actual time
      axis (s).  In row +0 of each scan, the same columns hold the diameter
      (nm) selected at each time step.  Row +1 ("Raw Data Counts") and
      row +2 ("Dead Time Counts") hold the corresponding count values.
"""

from __future__ import annotations

import numpy as np
from pathlib import Path
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

import openpyxl


# ── metadata keys that are useful for seeding the measurement model ──────────

# Keys in the instrument-level header block → hint name
_META_HINT_KEYS: Dict[str, str] = {
    'Reference Gas Temperature (K)': 'temperature_K',
    'Reference Gas Pressure (kPa)':  'pressure_kPa',
    'Tube Length (cm)':              'tube_length_cm',
    'Tube Diameter (cm)':            'tube_diameter_cm',
    'Channels/Decade':               'channels_per_decade',
    'Classifier HV Power Supply':    'hv_polarity',
}

# Keys in the per-scan metadata (first scan used) → hint name
_SCAN_HINT_KEYS: Dict[str, str] = {
    'Sheath Flow (L/min)':          'sheath_flow_Lmin',
    'Detector Inlet Flow (L/min)':  'aerosol_flow_Lmin',
    'Low Voltage (V)':              'v_min_V',
    'High Voltage (V)':             'v_max_V',
    'Lower Size (nm)':              'dp_min_nm',
    'Upper Size (nm)':              'dp_max_nm',
    'Impactor D50 (nm)':            'impactor_d50_nm',
}


# ── result dataclass ──────────────────────────────────────────────────────────

@dataclass
class TSIResult:
    """
    Parsed result from one sheet of a TSI SMPS Excel file.

    Attributes
    ----------
    instrument_meta : dict
        Header-block key/value pairs (instrument model, serial numbers,
        reference gas conditions, tube geometry, …).
    scan_datetimes : list
        Scan-start timestamps as returned by openpyxl (str or datetime).
    scan_meta_fields : list of str
        Column-header names for the per-scan metadata columns (cols 0 to
        dp_start − 1 in the header row).
    scan_meta : list of dict
        Per-scan metadata dicts (one dict per scan).
    diameters_dn : ndarray, shape (n_bins,)
        Diameter bin midpoints [nm] for the dN/dlogDp concentration grid.
    dn_matrix : ndarray, shape (n_scans, n_bins) or None
        dN/dlogDp concentration values; None when the block is empty.
    raw_times : ndarray, shape (n_times,) or None
        Time axis [s] for the raw-count block; None when absent.
    raw_counts : ndarray, shape (n_scans, n_times) or None
        Raw particle counts per scan per time step; None when absent.
    dead_time_counts : ndarray, shape (n_scans, n_times) or None
        Dead-time correction counts; None when absent.
    raw_diameters : ndarray, shape (n_scans, n_times) or None
        Diameter [nm] selected at each time step (from row +0); None when
        absent.
    """

    instrument_meta:  Dict[str, Any]
    scan_datetimes:   List[Any]
    scan_meta_fields: List[str]
    scan_meta:        List[Dict[str, Any]]
    diameters_dn:     np.ndarray
    dn_matrix:        Optional[np.ndarray]
    raw_times:        Optional[np.ndarray]
    raw_counts:       Optional[np.ndarray]
    dead_time_counts: Optional[np.ndarray]
    raw_diameters:    Optional[np.ndarray]

    @property
    def n_scans(self) -> int:
        return len(self.scan_datetimes)

    def instrument_hints(self) -> dict:
        """
        Return a flat dict of instrument parameter hints extracted from
        the metadata blocks.  Keys are chosen to match the spinbox names
        used in ModelTab so that seeding is straightforward.

        Returned keys (all optional, only present when parseable):
            temperature_K       – reference gas temperature [K]
            pressure_kPa        – reference gas pressure [kPa]
            pressure_Pa         – same, converted to Pa
            tube_length_cm      – DMA tube length [cm]
            tube_diameter_cm    – DMA tube diameter [cm]
            channels_per_decade – channels per decade of diameter
            hv_polarity         – 'negative' or 'positive' (lowercase string)
            sheath_flow_Lmin    – sheath flow rate [L/min]
            aerosol_flow_Lmin   – aerosol (sample) flow rate [L/min]
            v_min_V             – minimum scan voltage [V]
            v_max_V             – maximum scan voltage [V]
            dp_min_nm           – minimum selected diameter [nm]
            dp_max_nm           – maximum selected diameter [nm]
            impactor_d50_nm     – impactor 50 % cut diameter [nm]
            n_dp_bins           – number of dN/dlogDp diameter bins
        """
        hints: dict = {}

        # --- instrument-level metadata ---
        for src, dst in _META_HINT_KEYS.items():
            val = self.instrument_meta.get(src)
            if val is None:
                continue
            if dst == 'hv_polarity':
                hints[dst] = str(val).strip().lower()
            else:
                f = _try_float(val)
                if f is not None:
                    hints[dst] = f

        # Convenience: pressure also in Pa
        if 'pressure_kPa' in hints:
            hints['pressure_Pa'] = hints['pressure_kPa'] * 1000.0

        # --- first-scan metadata ---
        if self.scan_meta:
            sm = self.scan_meta[0]
            for src, dst in _SCAN_HINT_KEYS.items():
                val = sm.get(src)
                if val is not None:
                    f = _try_float(val)
                    if f is not None:
                        hints[dst] = f

        # --- number of diameter bins ---
        if self.diameters_dn is not None and len(self.diameters_dn) > 0:
            hints['n_dp_bins'] = len(self.diameters_dn)

        return hints

    def summary(self) -> str:
        """Human-readable summary of the parsed result."""
        lines = [
            f"Scans    : {self.n_scans}",
            f"Dp bins  : {len(self.diameters_dn)} "
            f"({self.diameters_dn[0]:.1f} – {self.diameters_dn[-1]:.1f} nm)"
            if len(self.diameters_dn) > 0 else "Dp bins  : 0",
        ]
        if self.dn_matrix is not None:
            lines.append(f"dN/dlogDp: {self.dn_matrix.shape}")
        if self.raw_counts is not None:
            lines.append(
                f"Raw cts  : {self.raw_counts.shape}  "
                f"(t = {self.raw_times[0]:.2f} – {self.raw_times[-1]:.2f} s)"
            )
        # Selected metadata values
        hints = self.instrument_hints()
        for k, label in [
            ('temperature_K',    'T ref'),
            ('pressure_kPa',     'P ref'),
            ('sheath_flow_Lmin', 'Qsh'),
            ('aerosol_flow_Lmin','Qa'),
            ('v_min_V',          'V_min'),
            ('v_max_V',          'V_max'),
        ]:
            if k in hints:
                lines.append(f"{label:8s} : {hints[k]}")
        return '\n'.join(lines)


# ── public API ────────────────────────────────────────────────────────────────

def list_sheets(filepath: str) -> List[str]:
    """Return the sheet names in a TSI SMPS Excel file."""
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def load_tsi_xlsx(filepath: str, sheet_name: Any = 0) -> TSIResult:
    """
    Parse one sheet from a TSI SMPS Excel export file.

    Parameters
    ----------
    filepath   : path to the .xlsx file
    sheet_name : sheet to read — name (str) or 0-based index (int)

    Returns
    -------
    TSIResult with all four blocks parsed.

    Raises
    ------
    FileNotFoundError  – file does not exist
    ValueError         – file does not look like a TSI SMPS export
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    # ── 1. Locate the scan header row ────────────────────────────────────
    header_row_idx: Optional[int] = None
    for idx, row in enumerate(rows):
        if row and row[0] == 'Scan Number':
            header_row_idx = idx
            break
    if header_row_idx is None:
        raise ValueError(
            "Could not find 'Scan Number' in column A. "
            "Please verify this is a TSI SMPS export file."
        )

    # ── 2. Instrument metadata block ─────────────────────────────────────
    instrument_meta: Dict[str, Any] = {}
    for row in rows[:header_row_idx]:
        if row and row[0] is not None and str(row[0]).strip():
            instrument_meta[str(row[0]).strip()] = row[1] if len(row) > 1 else None

    # ── 3. Decode the header row ─────────────────────────────────────────
    header = rows[header_row_idx]

    # dp_start: index of the first column that holds a float diameter value
    dp_start: Optional[int] = None
    for i, v in enumerate(header):
        if isinstance(v, float) and not isinstance(v, bool):
            dp_start = i
            break
    if dp_start is None:
        raise ValueError("Could not locate dN/dlogDp diameter columns in header row.")

    # none_sep_col: first None after dp_start (the separator before raw block)
    none_sep_col = len(header)
    for i in range(dp_start, len(header)):
        if header[i] is None:
            none_sep_col = i
            break

    # Diameter midpoints for dN/dlogDp
    diameters_dn = np.array(
        [header[i] for i in range(dp_start, none_sep_col) if header[i] is not None],
        dtype=float,
    )
    n_dp = len(diameters_dn)

    # raw_label_col: column whose header is "Raw Data - Time (s)"
    raw_label_col: Optional[int] = None
    for i in range(none_sep_col + 1, len(header)):
        if header[i] == 'Raw Data - Time (s)':
            raw_label_col = i
            break

    raw_time_start: Optional[int] = None
    raw_times: Optional[np.ndarray] = None
    n_raw = 0
    if raw_label_col is not None:
        raw_time_start = raw_label_col + 1
        t_vals = [header[i] for i in range(raw_time_start, len(header))
                  if header[i] is not None]
        raw_times = np.array(t_vals, dtype=float)
        n_raw = len(raw_times)

    # Per-scan metadata field names (string columns 0..dp_start-1)
    scan_meta_fields = [
        str(header[i]) if header[i] is not None else f'col{i}'
        for i in range(dp_start)
    ]

    # ── 4. Parse scan blocks ─────────────────────────────────────────────
    scan_meta:      List[Dict[str, Any]] = []
    scan_datetimes: List[Any] = []
    dn_rows:        List[List[float]] = []
    raw_count_rows: List[List[float]] = []
    dead_time_rows: List[List[float]] = []
    raw_diam_rows:  List[List[float]] = []

    i = header_row_idx + 1
    while i < len(rows):
        row0 = rows[i]
        if not _is_scan_row(row0):
            i += 1
            continue

        # Metadata dict for this scan
        sm: Dict[str, Any] = {}
        for j, fname in enumerate(scan_meta_fields):
            sm[fname] = row0[j] if j < len(row0) else None
        scan_meta.append(sm)
        scan_datetimes.append(row0[1] if len(row0) > 1 else None)

        # dN/dlogDp values (cols dp_start..none_sep_col-1)
        dn_vals = [_to_float(row0[j]) for j in range(dp_start, none_sep_col)
                   if j < len(row0)]
        pad = n_dp - len(dn_vals)
        dn_rows.append(dn_vals + [np.nan] * pad if pad > 0 else dn_vals[:n_dp])

        # Raw data: diameters in the same row, counts and dead-time in +1 / +2
        if raw_time_start is not None and n_raw > 0:
            rd = [_to_float(row0[j]) if j < len(row0) else np.nan
                  for j in range(raw_time_start, raw_time_start + n_raw)]
            raw_diam_rows.append(rd)

            row1 = rows[i + 1] if i + 1 < len(rows) else None
            row2 = rows[i + 2] if i + 2 < len(rows) else None

            if (row1 is not None and raw_label_col < len(row1)
                    and row1[raw_label_col] == 'Raw Data Counts'):
                rc = [_to_float(row1[j]) if j < len(row1) else 0.0
                      for j in range(raw_time_start, raw_time_start + n_raw)]
                raw_count_rows.append(rc)
            else:
                raw_count_rows.append([0.0] * n_raw)

            if (row2 is not None and raw_label_col < len(row2)
                    and row2[raw_label_col] == 'Dead Time Counts'):
                dt = [_to_float(row2[j]) if j < len(row2) else 0.0
                      for j in range(raw_time_start, raw_time_start + n_raw)]
                dead_time_rows.append(dt)
            else:
                dead_time_rows.append([0.0] * n_raw)

        i += 3  # advance past the 3-row block

    # ── 5. Assemble arrays ───────────────────────────────────────────────
    dn_matrix: Optional[np.ndarray] = None
    if dn_rows:
        arr = np.array(dn_rows, dtype=float)
        if not np.all(np.isnan(arr)):
            dn_matrix = arr

    raw_counts: Optional[np.ndarray] = None
    dead_time_counts: Optional[np.ndarray] = None
    raw_diameters: Optional[np.ndarray] = None

    if raw_count_rows:
        rc_arr = np.array(raw_count_rows, dtype=float)
        if not np.all(rc_arr == 0):
            raw_counts = rc_arr
            dead_time_counts = np.array(dead_time_rows, dtype=float)

    if raw_diam_rows:
        rd_arr = np.array(raw_diam_rows, dtype=float)
        if not np.all(np.isnan(rd_arr)):
            raw_diameters = rd_arr

    return TSIResult(
        instrument_meta  = instrument_meta,
        scan_datetimes   = scan_datetimes,
        scan_meta_fields = scan_meta_fields,
        scan_meta        = scan_meta,
        diameters_dn     = diameters_dn,
        dn_matrix        = dn_matrix,
        raw_times        = raw_times,
        raw_counts       = raw_counts,
        dead_time_counts = dead_time_counts,
        raw_diameters    = raw_diameters,
    )


# ── internal helpers ──────────────────────────────────────────────────────────

def _is_scan_row(row: tuple) -> bool:
    """True when row[0] holds a numeric scan index (integer value)."""
    if not row:
        return False
    v = row[0]
    if v is None or isinstance(v, bool):
        return False
    if isinstance(v, int):
        return True
    if isinstance(v, float):
        return v == int(v) and not np.isnan(v)
    return False


def _to_float(v: Any) -> float:
    """Cell value → float.  Returns np.nan on failure."""
    if v is None:
        return np.nan
    try:
        return float(v)
    except (TypeError, ValueError):
        return np.nan


def _try_float(v: Any) -> Optional[float]:
    """Cell value → float or None (instead of NaN) on failure."""
    if v is None:
        return None
    try:
        f = float(str(v).replace(',', '.'))
        return f if not np.isnan(f) else None
    except (TypeError, ValueError):
        return None
